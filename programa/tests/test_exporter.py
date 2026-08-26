import io
import pandas as pd
import openpyxl
from exporter import to_excel


def _weights_df() -> pd.DataFrame:
    return pd.DataFrame({
        "Ticker": ["AAPL", "MSFT", "GOOGL"],
        "Peso Óptimo (%)": ["50.00%", "30.00%", "20.00%"],
        "Retorno Esperado (%)": ["21.00%", "18.00%", "17.00%"],
        "Volatilidad (%)": ["22.00%", "19.00%", "20.00%"],
        "Contrib. Riesgo (%)": ["48.00%", "32.00%", "20.00%"],
    })


def _metrics() -> dict:
    return {
        "sharpe": 1.42,
        "annual_return": 0.183,
        "annual_vol": 0.128,
        "rf_rate": 0.0525,
        "horizon": "1 Mes",
    }


def test_to_excel_returns_bytes():
    result = to_excel(_weights_df(), _metrics())
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_to_excel_has_pesos_sheet():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    assert "Pesos" in wb.sheetnames


def test_to_excel_has_metricas_sheet():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    assert "Métricas" in wb.sheetnames


def test_to_excel_pesos_sheet_row_count():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    ws = wb["Pesos"]
    # 1 header + 3 data rows
    assert ws.max_row == 4


def test_to_excel_pesos_first_column_header():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    ws = wb["Pesos"]
    assert ws.cell(1, 1).value == "Ticker"


# ── Reported metrics must not hide the validation result ──────────────────────

from exporter import kpi_rows


def _validated_metrics() -> dict:
    return {**_metrics(), "oos_sharpe": 0.61, "oos_equal_weight_sharpe": 0.74, "oos_windows": 9}


def _labels(metrics: dict) -> list[str]:
    return [label for label, _ in kpi_rows(metrics)]


def test_kpi_rows_label_the_sharpe_as_in_sample():
    assert any("muestra" in label.lower() for label in _labels(_validated_metrics()))


def test_kpi_rows_include_the_out_of_sample_sharpe():
    rows = dict(kpi_rows(_validated_metrics()))
    assert any("0.61" in value for value in rows.values())


def test_kpi_rows_include_the_equal_weight_benchmark():
    rows = dict(kpi_rows(_validated_metrics()))
    assert any("0.74" in value for value in rows.values())


def test_kpi_rows_say_so_when_validation_did_not_run():
    rows = dict(kpi_rows(_metrics()))
    assert any("no disponible" in value.lower() for value in rows.values())


def test_kpi_rows_accept_a_metrics_dict_without_the_new_fields():
    assert len(kpi_rows(_metrics())) > 0


def test_kpi_rows_name_the_strategy_that_produced_the_weights():
    rows = dict(kpi_rows({**_validated_metrics(), "strategy": "Paridad de riesgo (ERC)"}))
    assert any("Paridad de riesgo" in v for v in rows.values())


def test_to_pdf_returns_bytes_for_validated_metrics():
    from exporter import to_pdf
    result = to_pdf(_weights_df(), _validated_metrics(), [])
    assert isinstance(result, bytes)
    assert len(result) > 0
