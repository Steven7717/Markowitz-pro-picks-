import io
import pandas as pd
import openpyxl
from exporter import to_excel


def _weights_df() -> pd.DataFrame:
    return pd.DataFrame({
        "Ticker": ["AAPL", "MSFT", "GOOGL"],
        "Peso Óptimo (%)": ["50.00%", "30.00%", "20.00%"],
        "Retorno Esperado (%)": ["21.00%", "18.00%", "17.00%"],
        "Volatilidad (%)": ["22.00%", "19.00%", "20.00%"],
        "Contrib. Riesgo (%)": ["48.00%", "32.00%", "20.00%"],
    })


def _metrics() -> dict:
    return {
        "sharpe": 1.42,
        "annual_return": 0.183,
        "annual_vol": 0.128,
        "rf_rate": 0.0525,
        "horizon": "1 Mes",
    }


def test_to_excel_returns_bytes():
    result = to_excel(_weights_df(), _metrics())
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_to_excel_has_pesos_sheet():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    assert "Pesos" in wb.sheetnames


def test_to_excel_has_metricas_sheet():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    assert "Métricas" in wb.sheetnames


def test_to_excel_pesos_sheet_row_count():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    ws = wb["Pesos"]
    # 1 header + 3 data rows
    assert ws.max_row == 4


def test_to_excel_pesos_first_column_header():
    wb = openpyxl.load_workbook(io.BytesIO(to_excel(_weights_df(), _metrics())))
    ws = wb["Pesos"]
    assert ws.cell(1, 1).value == "Ticker"
